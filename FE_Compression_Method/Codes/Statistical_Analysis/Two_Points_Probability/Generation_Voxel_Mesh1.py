#
from abaqus import *
from abaqusConstants import *
from caeModules import *
from driverUtils import executeOnCaeStartup
import string
import random
import numpy as np
import itertools
import openpyxl
#
from GeometryShape2D import *
#
executeOnCaeStartup()
Mdb()
#
#------------------------------------------------------------------------------------
#
session.journalOptions.setValues(replayGeometry=INDEX, recoverGeometry=INDEX)
#
Filenum = 1
#
for ifl in range(Filenum):
    #
    PartName, Pathname, IncId, FilId, strname = 'Part-1', 'Mesh-1.cae', 1, ifl+1, 'Circle'
    RLength, NumEle1, NumEle2, Ampf = 200.0, int(200*1.732), 200, 1.0
    f1 = 'C:\\Temp\\Midpoints2.txt'
    f2 = 'C:\\Temp\\NewOriangles1.txt'
    #
    UserVariables = {}
    UserVariables.update({'Type':'DEFORMABLE_BODY'})
    #
    #-----------------------------------------Circles------------------------------------------
    #
    if IncId == 1:
        IncSize = 6.16
        UserVariables.update({'Radius':IncSize*Ampf}) 
        ILength = UserVariables['Radius']
        Object = GeometryShape2D(UserVariables)
        sp = Object.Circles(PartName)
    #
    #-----------------------------------------Lobular2------------------------------------------
    #
    elif IncId == 2:
        IncSize = 4.2
        UserVariables.update({'Radius':IncSize*Ampf}) 
        ILength = 2.0*UserVariables['Radius']
        Object = GeometryShape2D(UserVariables)
        sp = Object.Lobular2s(PartName)
    #
    #-----------------------------------------Lobular3------------------------------------------
    #
    elif IncId == 3:
        IncSize = 3.4
        UserVariables.update({'Radius':IncSize*Ampf}) 
        ILength = (2.0/sqrt(3)+1.0)*UserVariables['Radius']
        Object = GeometryShape2D(UserVariables)
        sp = Object.Lobular3s(PartName)
    #
    #-----------------------------------------Lobular4------------------------------------------
    #
    elif IncId == 4:
        IncSize = 2.85
        UserVariables.update({'Radius':IncSize*Ampf}) 
        ILength = (1+sqrt(2))*UserVariables['Radius']
        Object = GeometryShape2D(UserVariables)
        sp = Object.Lobular4s(PartName)
    #
    #-----------------------------------------Spolygon3-----------------------------------------
    #
    elif IncId == 5:
        IncSize = [17.0, 2.25]
        UserVariables.update({'Slength':IncSize[0]*Ampf, 'Radius':IncSize[1]*Ampf}) 
        ILength = sqrt(3)/3.0*UserVariables['Slength']
        Object = GeometryShape2D(UserVariables)
        sp = Object.Spolygon3s(PartName)
    #
    #-----------------------------------------Spolygon4-----------------------------------------
    #
    elif IncId == 6:
        IncSize = [11.0, 2.75]
        UserVariables.update({'Slength':IncSize[0]*Ampf, 'Radius':IncSize[1]*Ampf}) 
        ILength = sqrt(2)/2.0*UserVariables['Slength']
        Object = GeometryShape2D(UserVariables)
        sp = Object.Spolygon4s(PartName)
    #
    #------------------------------------------Ellipses------------------------------------------
    #
    elif IncId == 7:
        IncSize = 8.5
        UserVariables.update({'Axis1':IncSize*Ampf, 'Axis2':IncSize/2.0*Ampf}) 
        ILength = UserVariables['Axis1']
        Object = GeometryShape2D(UserVariables)
        sp = Object.Ellipses(PartName)
    #
    #-------------------------------------------Kidneys-------------------------------------------
    #
    elif IncId == 8:
        IncSize = [5.6, 8.3, 3.8, 5.8]
        UserVariables.update({'R1':IncSize[0]*Ampf, 'R2':IncSize[1]*Ampf, 'K1':IncSize[2]*Ampf, 'a':IncSize[3]*Ampf}) 
        ILength = UserVariables['R2']
        Object = GeometryShape2D(UserVariables)
        sp = Object.Kidneys(PartName)
    #
    #-------------------------------Five-stars----------------------------------
    #
    else:
        IncSize = [8.6, 2.2]
        UserVariables.update({'Alength':IncSize[0]*Ampf, 'Radius':IncSize[1]*Ampf}) 
        ILength = UserVariables['Alength']*(cos(18.0/180*pi) + sin(18.0/180*pi)*tan(54.0/180*pi))
        Object = GeometryShape2D(UserVariables)
        sp = Object.F5stars(PartName)
    #
    #------------------------------------------------------------------------------------
    #
    myfile = open(f1,'r')
    lines = myfile.readlines()
    TranVec = []
    for line in lines:
        ele = string.split(line,',')
        TranVec.append((float(ele[0]),float(ele[1])))
    #
    myfile = open(f2,'r')
    lines = myfile.readlines()
    RoAngle = []
    for line in lines:
        ele = string.split(line,',')
        RoAngle.append((float(ele[0])))
    #
    #------------------------------------------------------------------------------------
    #
    instanceList = ()
    for inc in range(len(TranVec)):
        a = mdb.models['Model-1'].rootAssembly
        p = mdb.models['Model-1'].parts['Part-1']
        InsName = 'Part-1-' + str(inc+1)
        a.Instance(name=InsName, part=p, dependent=OFF)
        a.rotate(instanceList=(InsName, ), axisPoint=(0.0, 0.0, 0.0), axisDirection=(0.0, 0.0, 1.0), angle=RoAngle[inc])
        a.translate(instanceList=(InsName, ), vector=(TranVec[inc][0], TranVec[inc][1], 0.0))	
        #a.translate(instanceList=(InsName, ), vector=(RLength/2.0, RLength/2.0, 0.0))	
        instanceList = instanceList + ((a.instances[InsName], ))
    #
    a = mdb.models['Model-1'].rootAssembly
    a.InstanceFromBooleanMerge(name='Part-2', instances=instanceList, originalInstances=DELETE, domain=GEOMETRY)   
    del a.instances['Part-2-1'] 
    del mdb.models['Model-1'].parts['Part-1']
    mdb.models['Model-1'].parts.changeKey(fromName='Part-2', toName='Part-1')
    #
    #------------------------------------------------------------------------------------
    #
    a = mdb.models['Model-1'].rootAssembly
    p = mdb.models['Model-1'].parts['Part-1']
    a.Instance(name='Part-1-1', part=p, dependent=ON)
    f = a.instances['Part-1-1'].faces
    #
    VoxId = np.zeros((NumEle1, NumEle2), dtype=np.int)
    Xline = np.linspace(0 + RLength/(2*NumEle1), RLength - RLength/(2*NumEle1), NumEle1)
    Yline = np.linspace(0 + RLength/(2*NumEle2), RLength - RLength/(2*NumEle2), NumEle2)
    for x, y in itertools.product(range(NumEle1), range(NumEle2)):
        CenCrds = ((Xline[x], Yline[y], 0), )
        VoxId[x,y] = len(f.findAt(CenCrds))
    #
    #-------------------------------------------------------------------------------------------------------
    #
    f = open('C:\\Temp\\Statistical_Analysis\\Two_Points_Probability\\H6RVE.txt', 'w')
    for x, y in itertools.product(range(NumEle1), range(NumEle2)):
        if y < NumEle2 - 1:
            f.write(str(VoxId[x,y]) + ',')
        else:
            f.write(str(VoxId[x,y]) + '\n')
    f.close()
##
##-------------------------------------------------------------------------------------------------------
##
#s1 = mdb.models['Model-1'].ConstrainedSketch(name='__profile__', sheetSize=200.0)
#g, v, d, c = s1.geometry, s1.vertices, s1.dimensions, s1.constraints
#s1.setPrimaryObject(option=STANDALONE)
#s1.rectangle(point1=(0.0, 0.0), point2=(RLength, RLength))
#p = mdb.models['Model-1'].Part(name='Part-2', dimensionality=TWO_D_PLANAR, type=DEFORMABLE_BODY)
#p = mdb.models['Model-1'].parts['Part-2']
#p.BaseShell(sketch=s1)
#s1.unsetPrimaryObject()
#del mdb.models['Model-1'].sketches['__profile__']
##
#p = mdb.models['Model-1'].parts['Part-2']
#p.seedPart(size=RLength/NumEle1, deviationFactor=0.1, minSizeFactor=0.1)
##
#pickedRegions = p.faces
#p.setMeshControls(regions=pickedRegions, elemShape=QUAD, technique=STRUCTURED)
##
#elemType1 = mesh.ElemType(elemCode=CPE4R, elemLibrary=STANDARD, secondOrderAccuracy=OFF, hourglassControl=DEFAULT, 
#    distortionControl=DEFAULT)
#elemType2 = mesh.ElemType(elemCode=CPE3, elemLibrary=STANDARD)
#pickedRegions =(p.faces, )
#p.setElementType(regions=pickedRegions, elemTypes=(elemType1, elemType2))
##
#p.generateMesh()
##
##-------------------------------------------------------------------------------------------------------
##
#p = mdb.models['Model-1'].parts['Part-2']
#Set1Eles, Set2Eles = p.elements[0:0], p.elements[0:0]
##
#for x,y in itertools.product(range(NumEle1), repeat=2):
#    if VoxId[x,y] != 0:
#        Set1Eles += p.elements[x+y*NumEle1:x+y*NumEle1+1]
#    else:
#        Set2Eles += p.elements[x+y*NumEle1:x+y*NumEle1+1]
##
#p = mdb.models['Model-1'].parts['Part-2']
#p.Set(elements=Set1Eles, name='Set-1')
#p.Set(elements=Set2Eles, name='Set-2')
##
##-------------------------------------------------------------------------------------------------------
##
#Coords = np.ones((NumEle1*NumEle2, 3))
#Xline = np.linspace(0 + RLength/(2*NumEle1), RLength - RLength/(2*NumEle1), NumEle1)
#Yline = np.linspace(0 + RLength/(2*NumEle2), RLength - RLength/(2*NumEle2), NumEle2)
##
#for x, y in itertools.product(range(NumEle1), range(NumEle2)):
#	Coords[x+y*NumEle1,0:3] = [Xline[x], Yline[y], VoxId[x,y]]
##
#File = openpyxl.Workbook()
#Sheet = File.create_sheet("Sheet-1")
#for irow in range(NumEle1*NumEle2):
#    for icol in range(3):
#        Sheet.cell(row=irow+1,column=icol+1,value=Coords[irow,icol])
#File.save("Voxel_Element_Circle.xlsx")
##
#Vf = 1.0*len(Set1Eles)/NumEle1**2
#
#-------------------------------------------------------------------------------------------------------
#
#mdb.saveAs(pathName=Pathname)