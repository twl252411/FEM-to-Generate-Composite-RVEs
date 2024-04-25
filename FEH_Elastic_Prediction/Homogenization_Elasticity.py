from abaqus import *
from abaqusConstants import *
from odbAccess import *
import openpyxl
import os
import job
import string
import numpy as np
#
Inctypes = ['Circle', 'Lobular2', 'Lobular3', 'Lobular4', 'Spolygon3', 'Spolygon4', 'Ellipse', 'Kidney', 'Star']
strname = Inctypes[8]
filnum, resultfile = 10, 'Homogenized_Elastic_Constants_'+strname+'_60.xlsx'
ElaSts, ElaMat, CplMat = np.zeros((5,filnum+1)), np.zeros((9,filnum+1)), np.zeros((9,filnum+1))
#
for irves in range(filnum):
    #
    odb = openOdb('Job-'+str(irves+1)+'.odb')
    S, E, VRVE, Sinv = np.zeros((3,3)), np.zeros((3,3)), np.zeros((3,1)), np.zeros((3,3))
    #
    for j in range(3):
        for i in range(2):
            #
            Stepname = 'Step-' + str(j+1)
            instance = odb.rootAssembly.instances['PART-'+str(i+1)+'-1']
            #
            sf = odb.steps[Stepname].frames[-1].fieldOutputs['S']
            f1 = sf.getSubset(region=instance, position=INTEGRATION_POINT)
            fv1 = f1.values
            #
            ef = odb.steps[Stepname].frames[-1].fieldOutputs['E']
            f0 = ef.getSubset(region=instance, position=INTEGRATION_POINT)
            fv0 = f0.values
            #
            ivolf = odb.steps[Stepname].frames[-1].fieldOutputs['IVOL']
            f2= ivolf.getSubset(region=instance, position=INTEGRATION_POINT)
            fv2 = f2.values
            #
            evolf = odb.steps[Stepname].frames[-1].fieldOutputs['EVOL']
            f3 = evolf.getSubset(region=instance, position=WHOLE_ELEMENT)
            fv3 = f3.values
            #
            for iv in range(len(fv1)): 
            	for ip in range(3):
                    S[j,ip] = S[j,ip] + fv1[iv].data[ip+ip//2]*fv2[iv].data
                    E[j,ip] = E[j,ip] + fv0[iv].data[ip+ip//2]*fv2[iv].data
            #
            for iv in range(len(fv3)): 
                VRVE[j,0] = VRVE[j,0] + fv3[iv].data
        #
        S[j,:], E[j,:] = S[j,:]/VRVE[j,0], E[j,:]/VRVE[j,0]
        Sinv[j,:] = E[j,:]/S[j,j]  
    #
    odb.close() 
    #
#    mu12, mu21 = -E[1,0]/E[0,0], -E[0,1]/E[1,1]
#    E1, E2, G = S[0,0]/E[0,0], S[1,1]/E[1,1], S[2,2]/E[2,2]
    
    mu12, mu21 = Sinv[1,0]/(Sinv[1,0]-Sinv[0,0]), Sinv[0,1]/(Sinv[0,1]-Sinv[1,1])
    E1, E2, G = (1.0-mu12**2)/Sinv[0,0], (1.0-mu21**2)/Sinv[1,1], 1.0/Sinv[2,2]
    #
    ElaSts[:,irves] = np.array([[E1], [E2], [G], [mu12], [mu21]])[:,0]
    ElaMat[:,irves] = np.linalg.inv(Sinv).reshape(9,1)[:,0]
    CplMat[:,irves] = Sinv.reshape(9,1)[:,0] 
#
ElaSts[:,filnum] = np.average(ElaSts[:,0:filnum], axis=1)
ElaMat[:,filnum] = np.average(ElaMat[:,0:filnum], axis=1)
CplMat[:,filnum] = np.average(CplMat[:,0:filnum], axis=1)
#
File = openpyxl.Workbook()
Sheet1, Sheet2, Sheet3 = File.create_sheet("sheet-1"), File.create_sheet("sheet-2"), File.create_sheet("sheet-3")
#
for irow in range(5):
    for icol in range(filnum+1): 
        Sheet1.cell(row=irow+1,column=icol+1,value=ElaSts[irow,icol])
for irow in range(9):
    for icol in range(filnum+1): 
        Sheet2.cell(row=irow+1,column=icol+1,value=ElaMat[irow,icol])
        Sheet3.cell(row=irow+1,column=icol+1,value=CplMat[irow,icol])
File.save(resultfile)

