# -*- coding: utf-8 -*-
"""
Created on Sat Feb  3 09:06:39 2024 @author: TIAN
"""
import string
import openpyxl
import pandas as pd
import numpy as np
#
#----------------------------- Open the sta file ------------------------------
#
resultfile = 'S11E11-Compression.xlsx'
File = openpyxl.Workbook()
#
for ifile in range(9,18):
    #
    fad = './S11-Compression'+str(ifile)+'.xlsx'
    datamat = pd.read_excel(io=fad, header=None, sheet_name=1, usecols=[0])
    #
    stafile = open('Job-'+str(ifile+1)+'.sta','r')
    lines = stafile.readlines()
    stafile.close()
    #
    emat = []
    emat.append([0, datamat[0][0]*1000000])
    #
    #-------------------- Assemble the nodes and elements ------------------------
    #
    tsum, count = 0, 0
    for iline in range(5, len(lines)-2):
        linstrs = lines[iline].split( )
        if linstrs[2][-1] != 'U':
            tsum = tsum + float(linstrs[8])
            count = count + 1
            emat.append([tsum,datamat[0][count]*1000000])
    #
    #-------------------- Save the new strain file --------------------------------
    #
    Sheet1 = File.create_sheet("Sheet-"+str(ifile+1))
    for irow in range(len(emat)):
        for icol in range(2):
            Sheet1.cell(row=irow+1, column=icol+1, value=emat[irow][icol])
    File.save(resultfile)

