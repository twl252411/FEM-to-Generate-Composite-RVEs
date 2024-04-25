# -*- coding: utf-8 -*-
"""
Created on Thu Jul 27 14:39:22 2023 @author: Tian
"""
#
import numpy as np
import matplotlib.pyplot as plt
import scipy.spatial as spt
from scipy import stats
import openpyxl
#
filenum, smothpt, strname = 10, 500, 'Circle'
filename = 'C:\\Temp\\Statistical_Analysis\\Nearest_Neighbour_Distance_Angle\\'+strname+'_RVE\\NNDA'+strname+'.xlsx'
nndpdf, nnacdf = np.zeros((smothpt+1,filenum*3+3)), np.zeros((361,filenum*2+2))
tmppdf, tmpcdf = np.zeros((smothpt+1,2)), np.zeros((361,1))
#
glomax, glomin = 0, 200.0
for ifile in range(filenum):
    #
    points = np.loadtxt('C:\\Temp\\Generated_Points_Angles\\' + strname + '\\NewCentroids'+str(ifile+1)+'.txt', delimiter=',')
    kndist = np.zeros((len(points), 2))
    #
    kt = spt.KDTree(data=points) 
    for i in range(len(points)):
        find_point = points[i,:] 
        d, id1 = kt.query(find_point,3)
        kndist[i,:] = d[1:3]
    #
    dmax, dmin = np.max(kndist[:,1]), np.min(kndist[:,0])
    if dmax > glomax: glomax = dmax
    if dmin < glomin: glomin = dmin
#
for ifile in range(filenum):
    #
    points = np.loadtxt('C:\\Temp\\Generated_Points_Angles\\' + strname + '\\NewCentroids'+str(ifile+1)+'.txt', delimiter=',')
    kndist = np.zeros((len(points), 2))
    knang = np.zeros((len(points), 1))
    #
    kt = spt.KDTree(data=points) 
    for i in range(len(points)):
        find_point = points[i,:] 
        d, id1 = kt.query(find_point,3)
        kndist[i,:] = d[1:3]
        vec = points[id1[1],:] - find_point
        xvec = np.array([1, 0]) 
        if vec[1] >= 0:
            knang[i] = np.arccos(np.dot(vec,xvec)/np.linalg.norm(vec))*180/np.pi
        else:
            knang[i] = 360 - np.arccos(np.dot(vec,xvec)/np.linalg.norm(vec))*180/np.pi
    #---------------------------------------
    ydata = np.zeros((smothpt+1,2))
    for i in range(2):
        data = kndist[:,i]
        gkde = stats.gaussian_kde(data, bw_method = 'scott')
        xdata = np.linspace(glomin, glomax, smothpt+1)
        ydata[:,i] = gkde.evaluate(xdata)
    #
    nndpdf[:,ifile*3+0], nndpdf[:,ifile*3+1], nndpdf[:,ifile*3+2] = xdata, ydata[:,0], ydata[:,1]
    tmppdf[:,0] += ydata[:,0]/filenum
    tmppdf[:,1] += ydata[:,1]/filenum
    #
    cuang = np.zeros((361,2))
    for i in range(361):
        cuang[i,0] = i
        cuang[i,1] = np.sum(knang <= i)/len(points)
    #
    nnacdf[:,ifile*2+0], nnacdf[:,ifile*2+1] = cuang[:,0], cuang[:,1]
    tmpcdf[:,0] += cuang[:,1]/filenum
    #
#
nndpdf[:,filenum*3+0] = nndpdf[:,0]
nndpdf[:,filenum*3+1], nndpdf[:,filenum*3+2] = tmppdf[:,0], tmppdf[:,1]
nnacdf[:,filenum*2+0], nnacdf[:,filenum*2+1] = nnacdf[:,0], tmpcdf[:,0]
#
File = openpyxl.Workbook()
#
Sheet = File.create_sheet("sheet-1")
for irow in range(smothpt+1):
    for icol in range(filenum*3+3):
        Sheet.cell(row=irow+1,column=icol+1,value=nndpdf[irow,icol])
#
Sheet1 = File.create_sheet("sheet-2")
for irow in range(361):
    for icol in range(filenum*2+2):
        Sheet1.cell(row=irow+1,column=icol+1,value=nnacdf[irow,icol])
#
File.save(filename)


