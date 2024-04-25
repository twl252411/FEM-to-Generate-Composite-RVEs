#
import numpy as np
import string
import random
import openpyxl
import pandas as pd
#
#------------------------------------------------------------------------------------
#
filnum, srad, typename = 10, 20, 'circle'
datamat = np.zeros((2,filnum+3))
#
for ifile in range(filnum):
    #
    fad = './Local_Volume_Fraction_'+typename+'_'+str(ifile+1)+'_'+str(srad)+'.xlsx'
    datamat[0,ifile] = np.max(pd.read_excel(io=fad, header=None, sheet_name=1, usecols=[2]))
    datamat[1,ifile] = np.min(pd.read_excel(io=fad, header=None, sheet_name=1, usecols=[2]))
    datamat[0,filnum] = datamat[0,filnum] + datamat[0,ifile]/filnum
    datamat[1,filnum] = datamat[1,filnum] + datamat[1,ifile]/filnum
    #
#
datamat[0,filnum+1], datamat[0,filnum+2] = np.max(datamat[0,0:filnum])-datamat[0,filnum], datamat[0,filnum]-np.min(datamat[0,0:filnum])
datamat[1,filnum+1], datamat[1,filnum+2] = np.max(datamat[1,0:filnum])-datamat[1,filnum], datamat[1,filnum]-np.min(datamat[1,0:filnum])
#
opfname = 'Local_Volume_Fraction_'+typename+'_'+str(srad)+'_1.xlsx'
File = openpyxl.Workbook()
Sheet = File.create_sheet("Sheet-1")
#
for irow in range(2):
    for icol in range(filnum+3):
        Sheet.cell(row=irow+1, column=icol+1, value=datamat[irow,icol])
File.save(opfname)
