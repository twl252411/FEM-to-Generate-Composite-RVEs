# -*- coding: utf-8 -*-
"""
Created on Fri Jan 19 22:44:34 2024

@author: Tian
"""
#
import numpy as np
import itertools
import openpyxl
#
#-------------------------------------------------------------------------------------------------------
#
NumEle, RLength = 51, 200.0
#
Coords = np.ones((NumEle, 4*NumEle))
Xline = np.linspace(0.5, RLength-0.5, NumEle)
#
for x in range(NumEle):
    Coords[:,x*2], Coords[:,x*2+1] = x*(RLength-1)/(NumEle-1), Xline[:]
    Coords[:,x*2+2*NumEle], Coords[:,x*2+1+2*NumEle] = Xline[:], x*(RLength-1)/(NumEle-1)
#
File = openpyxl.Workbook()
Sheet = File.create_sheet("Sheet-1")
for irow in range(NumEle):
    for icol in range(4*NumEle):
        Sheet.cell(row=irow+1,column=icol+1,value=Coords[irow,icol])
File.save("Voxel_Mesh.xlsx")