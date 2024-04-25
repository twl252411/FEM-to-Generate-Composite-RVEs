#
import numpy as np
import string
import random
import openpyxl
import pandas as pd
#
#------------------------------------------------------------------------------------
#
filnum, typename = 10, 'lobular2'
datamat = np.zeros((12,filnum+3))
#
for isize in range(6):
    #
    srad = 20*(isize+1)
    #
    for ifile in range(filnum):
        #
        fad = './'+str(srad)+'//Local_Volume_Fraction_'+typename+'_'+str(ifile+1)+'_'+str(srad)+'.xlsx'
        datamat[0+isize*2,ifile] = np.max(pd.read_excel(io=fad, header=None, sheet_name=1, usecols=[2]))
        datamat[1+isize*2,ifile] = np.min(pd.read_excel(io=fad, header=None, sheet_name=1, usecols=[2]))
        datamat[0+isize*2,filnum] = datamat[0+isize*2,filnum] + datamat[0+isize*2,ifile]/filnum
        datamat[1+isize*2,filnum] = datamat[1+isize*2,filnum] + datamat[1+isize*2,ifile]/filnum
        #
    #
    datamat[0+isize*2,filnum+1] = np.max(datamat[0+isize*2,0:filnum])-datamat[0+isize*2,filnum]
    datamat[0+isize*2,filnum+2] = datamat[0+isize*2,filnum]-np.min(datamat[0+isize*2,0:filnum])
    datamat[1+isize*2,filnum+1] = np.max(datamat[1+isize*2,0:filnum])-datamat[1+isize*2,filnum]
    datamat[1+isize*2,filnum+2] = datamat[1+isize*2,filnum]-np.min(datamat[1+isize*2,0:filnum])
#
opfname = 'Local_Volume_Fraction_'+typename+'_1.xlsx'
File = openpyxl.Workbook()
Sheet = File.create_sheet("Sheet-1")
#
for irow in range(12):
    for icol in range(filnum+3):
        if irow < 6:
            Sheet.cell(row=irow+1, column=icol+1, value=datamat[irow*2,icol])
        else:
            Sheet.cell(row=irow+1, column=icol+1, value=datamat[(irow-6)*2+1,icol])
File.save(opfname)
