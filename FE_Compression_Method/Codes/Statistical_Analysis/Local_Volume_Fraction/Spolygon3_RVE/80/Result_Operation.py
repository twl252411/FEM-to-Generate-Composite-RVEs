#
import numpy as np
import string
import random
import openpyxl
import pandas as pd
#
#------------------------------------------------------------------------------------
#
filnum, srad, typename = 10, 80, 'spolygon3'
datamat = np.zeros((1681,13))
#
for ifile in range(filnum):
    #
    fad = './Local_Volume_Fraction_'+typename+'_'+str(ifile+1)+'_'+str(srad)+'.xlsx'
    #
    if ifile == 0:
        datamat[:,0:3] = pd.read_excel(io=fad, header=None, sheet_name=1, usecols=[0,1,2])
        datamat[:,12] = datamat[:,12] + datamat[:,2]/filnum
    else:
        datamat[:,ifile+2:ifile+3] = pd.read_excel(io=fad, header=None, sheet_name=1, usecols=[2])
        datamat[:,12] = datamat[:,12] + datamat[:,ifile+2]/filnum
    #
opfname = 'Local_Volume_Fraction_'+typename+'_'+str(srad)+'.xlsx'
File = openpyxl.Workbook()
Sheet = File.create_sheet("Sheet-1")
#
for irow in range(1681):
    for icol in range(13):
        Sheet.cell(row=irow+1, column=icol+1, value=datamat[irow,icol])
File.save(opfname)
