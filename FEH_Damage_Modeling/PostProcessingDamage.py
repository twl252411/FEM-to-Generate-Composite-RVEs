from abaqus import *
from abaqusConstants import *
from odbAccess import *
import string
import openpyxl
#
#----------------------------------------------------------------------------------------------------
#
for ifile in range(18):
    #
    if ifile < 9:
    	resultfile = 'S11-Tension-'+str(ifile+1)+'.xlsx'
    else:
    	resultfile = 'S11-Compression-'+str(ifile+1)+'.xlsx'
    #
    odb = openOdb('Job-'+str(ifile+1)+'.odb')
    scratchOdb = session.ScratchOdb(odb)
    datum = scratchOdb.rootAssembly.DatumCsysByThreePoints(name='CSYS-1', coordSysType=CARTESIAN, origin=(0.0, 0.0, 0.0), 
        point1=(1.0, 0.0, 0.0), point2=(0.0, 1.0, 0.0))
    #
    numframe = len(odb.steps['Step-1'].frames)
    ES, VRVE = [[0.0 for col in range(2)] for row in range(numframe)], [0.0 for col in range(numframe)]
    #
    for j in range(numframe):
        #
        for i in range(2):
            #
            instance = odb.rootAssembly.instances['PART-'+str(i+1)+'-1']
            #
            stres = odb.steps['Step-1'].frames[j].fieldOutputs['S'].getTransformedField(datumCsys=datum)
            ivolf = odb.steps['Step-1'].frames[j].fieldOutputs['IVOL']
            evolf = odb.steps['Step-1'].frames[j].fieldOutputs['EVOL']
            #
            fld1 = stres.getSubset(region=instance, position=INTEGRATION_POINT).values
            fld3 = ivolf.getSubset(region=instance, position=INTEGRATION_POINT).values
            fld4 = evolf.getSubset(region=instance, position=WHOLE_ELEMENT).values
            #
            for v in range(len(fld1)):
                ES[j][0] = ES[j][0] + fld1[v].data[0]*fld3[v].data
            #
            for v in range(len(fld4)):
                VRVE[j] = VRVE[j] + fld4[v].data
        #
        ES[j][0] = (-1)**(ifile//9)*ES[j][0]/VRVE[j]
    #
    odb.close()
    #
    File = openpyxl.Workbook()
    Sheet1 = File.create_sheet("Sheet-"+str(ifile+1))
    for irow in range(numframe):
        for icol in range(1): 
            Sheet1.cell(row=irow+1, column=icol+1, value=ES[irow][icol])
    File.save(resultfile)