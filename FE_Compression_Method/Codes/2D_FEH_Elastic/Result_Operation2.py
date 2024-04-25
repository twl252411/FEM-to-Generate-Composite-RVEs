#
import numpy as np
import string
import random
import openpyxl
import pandas as pd
#
#------------------------------------------------------------------------------------
#
Inctypes = ['Circle', 'Lobular2', 'Lobular3', 'Lobular4', 'Spolygon3', 'Spolygon4', 'Ellipse', 'Kidney', 'Star']
filnum = 9
#
datamat = np.zeros((filnum,5*3))
for ifile in range(filnum):
    #
    fad = './Homogenized_Elastic_Constants_'+Inctypes[ifile]+'.xlsx'
    #
    mat = pd.read_excel(io=fad, header=None, sheet_name=3)
    tmpmat = np.zeros((5,11))
    for it in range(11):
        tmpmat[0,it], tmpmat[1,it], tmpmat[4,it] = 1/mat[it][0]*1000.0, 1/mat[it][4]*1000.0, 1/mat[it][8]*1000.0
        tmpmat[2,it], tmpmat[3,it] = -mat[it][1]/mat[it][0], -mat[it][3]/mat[it][4]
    # for it in range(5):
    #     tmpmat[it,10] = np.average(tmpmat[it,0:10])
    #
    for k in range(5):
        datamat[ifile,k*2+0] = tmpmat[k,10]
        datamat[ifile,k*2+1] = (np.max(tmpmat[k,:]) - np.min(tmpmat[k,:]))/2.0
#
opfname = 'Homogenized_Elastic_Constants1.xlsx'
File = openpyxl.Workbook()
Sheet = File.create_sheet("Sheet-1")
#
for irow in range(filnum):
    for icol in range(10):
        Sheet.cell(row=irow+1, column=icol+1, value=datamat[irow,icol])
File.save(opfname)
