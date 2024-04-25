#
from abaqus import *
from abaqusConstants import *
from caeModules import *
from driverUtils import executeOnCaeStartup
import numpy as np
import string
import random
import openpyxl
#
from GeometryShape2D import *
from GetXYCoordinates import *
#
executeOnCaeStartup()
Mdb()
#
#------------------------------------------------------------------------------------
#
filnum, RLength, Sradius, NRow, Ampf = 10, 200.0, 20, 40, 1.0
#
PartName, pathname, strname, IncId = 'Part-1', 'RVE_1.cae', 'Spolygon4', 6
#
ifile = 1
f1 = 'C:\\Temp\\Generated_Points_Angles\\'+strname+'\\NewCentroids'+str(ifile)+'.txt'
f2 = 'C:\\Temp\\Generated_Points_Angles\\'+strname+'\\NewOriangles'+str(ifile)+'.txt'
opfname = "C:\\Temp\\Statistical_Analysis\\Local_Volume_Fraction\\Spolygon4_RVE\\"+str(Sradius)+"\\Local_Volume_Fraction_"+strname+"_"+str(ifile)+"_"+str(Sradius)+".xlsx"
#
UserVariables = {}
UserVariables.update({'Type':'DEFORMABLE_BODY', 'dPid':2})
#
#-----------------------------------------Circles------------------------------------------
#
if IncId == 1:
    IncSize = 6.0
    UserVariables.update({'Radius':IncSize*Ampf}) 
    ILength = UserVariables['Radius']
    Object = GeometryShape2D(UserVariables)
    sp = Object.Circles(PartName)
    pa = GetXYCoordinates(UserVariables).Partition(PartName)
    Coord1 = GetXYCoordinates(UserVariables).GetCoords(PartName)
#
#-----------------------------------------Lobular2------------------------------------------
#
elif IncId == 2:
    IncSize = 4.2
    UserVariables.update({'Radius':IncSize*Ampf}) 
    ILength = 2.0*UserVariables['Radius']
    Object = GeometryShape2D(UserVariables)
    sp = Object.Lobular2s(PartName)
    pa = GetXYCoordinates(UserVariables).Partition(PartName)
    Coord1 = GetXYCoordinates(UserVariables).GetCoords(PartName)
#
#-----------------------------------------Lobular3------------------------------------------
#
elif IncId == 3:
    IncSize = 3.4
    UserVariables.update({'Radius':IncSize*Ampf}) 
    ILength = (2.0/sqrt(3)+1.0)*UserVariables['Radius']
    Object = GeometryShape2D(UserVariables)
    sp = Object.Lobular3s(PartName)
    pa = GetXYCoordinates(UserVariables).Partition(PartName)
    Coord1 = GetXYCoordinates(UserVariables).GetCoords(PartName)
#
#-----------------------------------------Lobular4------------------------------------------
#
elif IncId == 4:
    IncSize = 2.85
    UserVariables.update({'Radius':IncSize*Ampf}) 
    ILength = (1+sqrt(2))*UserVariables['Radius']
    Object = GeometryShape2D(UserVariables)
    sp = Object.Lobular4s(PartName)
    pa = GetXYCoordinates(UserVariables).Partition(PartName)
    Coord1 = GetXYCoordinates(UserVariables).GetCoords(PartName)
#
#-----------------------------------------Spolygon3-----------------------------------------
#
elif IncId == 5:
    IncSize = [17.0, 2.25]
    UserVariables.update({'Slength':IncSize[0]*Ampf, 'Radius':IncSize[1]*Ampf}) 
    ILength = sqrt(3)/3.0*UserVariables['Slength']
    Object = GeometryShape2D(UserVariables)
    sp = Object.Spolygon3s(PartName)
    pa = GetXYCoordinates(UserVariables).Partition(PartName)
    Coord1 = GetXYCoordinates(UserVariables).GetCoords(PartName)
#
#-----------------------------------------Spolygon4-----------------------------------------
#
elif IncId == 6:
    IncSize = [11.0, 2.75]
    UserVariables.update({'Slength':IncSize[0]*Ampf, 'Radius':IncSize[1]*Ampf}) 
    ILength = sqrt(2)/2.0*UserVariables['Slength']
    Object = GeometryShape2D(UserVariables)
    sp = Object.Spolygon4s(PartName)
    pa = GetXYCoordinates(UserVariables).Partition(PartName)
    Coord1 = GetXYCoordinates(UserVariables).GetCoords(PartName)
#
#------------------------------------------Ellipses------------------------------------------
#
elif IncId == 7:
    IncSize = 8.5
    UserVariables.update({'Axis1':IncSize*Ampf, 'Axis2':IncSize/2.0*Ampf}) 
    ILength = UserVariables['Axis1']
    Object = GeometryShape2D(UserVariables)
    sp = Object.Ellipses(PartName)
    pa = GetXYCoordinates(UserVariables).Partition(PartName)
    Coord1 = GetXYCoordinates(UserVariables).GetCoords(PartName)
#
#-------------------------------------------Kidneys-------------------------------------------
#
elif IncId == 8:
    IncSize = [5.6, 8.3, 3.8, 5.8]
    UserVariables.update({'R1':IncSize[0]*Ampf, 'R2':IncSize[1]*Ampf, 'K1':IncSize[2]*Ampf, 'a':IncSize[3]*Ampf}) 
    ILength = UserVariables['R2']
    Object = GeometryShape2D(UserVariables)
    sp = Object.Kidneys(PartName)
    pa = GetXYCoordinates(UserVariables).Partition(PartName)
    Coord1 = GetXYCoordinates(UserVariables).GetCoords(PartName)
#
#-------------------------------Five-stars----------------------------------
#
else:
    IncSize = [8.6, 2.2]
    UserVariables.update({'Alength':IncSize[0]*Ampf, 'Radius':IncSize[1]*Ampf}) 
    ILength = UserVariables['Alength']*(cos(18.0/180*pi) + sin(18.0/180*pi)*tan(54.0/180*pi))
    Object = GeometryShape2D(UserVariables)
    sp = Object.F5stars(PartName)
    pa = GetXYCoordinates(UserVariables).Partition(PartName)
    Coord1 = GetXYCoordinates(UserVariables).GetCoords(PartName)
#
#------------------------------------------------------------------------------------
#
myfile = open(f1,'r')
lines = myfile.readlines()
TranVec = []
for line in lines:
    ele = string.split(line,',')
    TranVec.append((float(ele[0]),float(ele[1])))
#
myfile = open(f2,'r')
lines = myfile.readlines()
RoAngle = []
for line in lines:
    ele = string.split(line,',')
    RoAngle.append((float(ele[0])))
#
#------------------------------------------------------------------------------------
#
instanceList = ()
for inc in range(len(TranVec)):
    a = mdb.models['Model-1'].rootAssembly
    p = mdb.models['Model-1'].parts['Part-1']
    InsName = 'Part-1-' + str(inc+1)
    a.Instance(name=InsName, part=p, dependent=OFF)
    a.rotate(instanceList=(InsName, ), axisPoint=(0.0, 0.0, 0.0), axisDirection=(0.0, 0.0, 1.0), angle=RoAngle[inc])
    a.translate(instanceList=(InsName, ), vector=(TranVec[inc][0], TranVec[inc][1], 0.0))	
    a.translate(instanceList=(InsName, ), vector=(RLength/2.0, RLength/2.0, 0.0))	
    instanceList = instanceList + ((a.instances[InsName], ))
#
a = mdb.models['Model-1'].rootAssembly
a.InstanceFromBooleanMerge(name='Part-2', instances=instanceList, originalInstances=DELETE, domain=GEOMETRY)   
a.LinearInstancePattern(instanceList=('Part-2-1', ), direction1=(1.0, 0.0, 0.0), direction2=(0.0, 1.0, 0.0), 
    number1=3, number2=3, spacing1=RLength, spacing2=RLength)
#
a = mdb.models['Model-1'].rootAssembly
a.InstanceFromBooleanMerge(name='Part-3', instances=(a.instances['Part-2-1-lin-3-1'], a.instances['Part-2-1-lin-3-2'], 
    a.instances['Part-2-1-lin-3-3'], a.instances['Part-2-1-lin-2-3'], a.instances['Part-2-1-lin-2-2'], a.instances['Part-2-1-lin-2-1'], 
    a.instances['Part-2-1'], a.instances['Part-2-1-lin-1-2'], a.instances['Part-2-1-lin-1-3'], ), originalInstances=DELETE, domain=GEOMETRY)
a.deleteFeatures(('Part-3-1',))
#    
del mdb.models['Model-1'].parts['Part-1']
del mdb.models['Model-1'].parts['Part-2']
mdb.models['Model-1'].parts.changeKey(fromName='Part-3', toName='Part-1')
#
#------------------------------------------------------------------------------------
#
s = mdb.models['Model-1'].ConstrainedSketch(name='__profile__', sheetSize=200.0)
g, v, d, c = s.geometry, s.vertices, s.dimensions, s.constraints
s.setPrimaryObject(option=STANDALONE)
s.CircleByCenterPerimeter(center=(0.0, 0.0), point1=(Sradius, 0.0))
p = mdb.models['Model-1'].Part(name='Part-2', dimensionality=TWO_D_PLANAR, type=DEFORMABLE_BODY)
p = mdb.models['Model-1'].parts['Part-2']
p.BaseShell(sketch=s)
s.unsetPrimaryObject()
del mdb.models['Model-1'].sketches['__profile__']
#
dic = mdb.models['Model-1'].parts['Part-2'].getMassProperties( )
volum2 = dic['area']
#
#------------------------------------------------------------------------------------
#
LocalVF = np.zeros(((NRow + 1)**2, 3)) 
# 
for i in range(NRow + 1):
    for j in range(NRow + 1):
        #
        xcoord, ycoord = i*RLength/NRow - 0.5*RLength, j*RLength/NRow - 0.5*RLength  
        #
        a = mdb.models['Model-1'].rootAssembly
        p = mdb.models['Model-1'].parts['Part-2']
        a.Instance(name='Part-2-1', part=p, dependent=OFF)
        a.translate(instanceList=('Part-2-1', ), vector=(xcoord, ycoord, 0.0))
        #
        a = mdb.models['Model-1'].rootAssembly
        p = mdb.models['Model-1'].parts['Part-1']
        a.Instance(name='Part-1-1', part=p, dependent=OFF)
        a.translate(instanceList=('Part-1-1', ), vector=(-RLength, -RLength, 0.0))
        #
        a = mdb.models['Model-1'].rootAssembly
        a.InstanceFromBooleanCut(name='Part-3', instanceToBeCut=a.instances['Part-2-1'], 
            cuttingInstances=(a.instances['Part-1-1'], ), originalInstances=DELETE)
        a.deleteFeatures(('Part-3-1',))
        #
        dic = mdb.models['Model-1'].parts['Part-3'].getMassProperties( )
        volum3 = dic['area']
        #
        LocalVF[i*(NRow+1)+j,:] = [xcoord, ycoord, 1.0-volum3/volum2]
        #
        del mdb.models['Model-1'].parts['Part-3']
#
File = openpyxl.Workbook()
Sheet = File.create_sheet("sheet-1")
#
for irow in range((NRow + 1)**2):
    for icol in range(3):
        Sheet.cell(row=irow+1,column=icol+1,value=LocalVF[irow,icol])
#
File.save(opfname)